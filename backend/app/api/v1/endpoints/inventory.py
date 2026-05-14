from io import BytesIO

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.deps import CurrentUser, DbDep
from app.crud import inventory as inventory_crud
from app.schemas.inventory import MonthlyReport, SalespersonReport, StockRow

router = APIRouter(prefix="/inventory", tags=["inventory"])


@router.get("/stock", response_model=list[StockRow])
def list_stock(
    db: DbDep,
    _: CurrentUser,
    search: str | None = None,
    category_id: int | None = None,
    low_only: bool = False,
    include_inactive: bool = False,
) -> list[StockRow]:
    return inventory_crud.list_stock(
        db,
        search=search,
        category_id=category_id,
        low_only=low_only,
        include_inactive=include_inactive,
    )


@router.get("/monthly-report", response_model=MonthlyReport)
def get_monthly_report(
    year: int,
    month: int,
    db: DbDep,
    _: CurrentUser,
) -> MonthlyReport:
    try:
        return inventory_crud.monthly_report(db, year=year, month=month)
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc


@router.get("/monthly-report.xlsx")
def export_monthly_report(
    year: int,
    month: int,
    db: DbDep,
    _: CurrentUser,
) -> StreamingResponse:
    try:
        report = inventory_crud.monthly_report(db, year=year, month=month)
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    title = f"進銷存月報 {year}-{month:02d}"
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    headers = [
        "SKU",
        "商品名稱",
        "分類",
        "期初庫存",
        "本月進貨",
        "本月銷貨",
        "盤點異動",
        "期末庫存",
        "進貨金額",
        "銷貨金額",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1976D2")
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(report.rows, start=4):
        ws.cell(row=idx, column=1, value=row.sku)
        ws.cell(row=idx, column=2, value=row.name)
        ws.cell(row=idx, column=3, value=row.category_name or "")
        ws.cell(row=idx, column=4, value=row.opening_stock)
        ws.cell(row=idx, column=5, value=row.qty_in)
        ws.cell(row=idx, column=6, value=row.qty_out)
        ws.cell(row=idx, column=7, value=row.adjustment)
        ws.cell(row=idx, column=8, value=row.closing_stock)
        ws.cell(row=idx, column=9, value=float(row.purchase_amount))
        ws.cell(row=idx, column=10, value=float(row.sales_amount))
        ws.cell(row=idx, column=9).number_format = "#,##0.00"
        ws.cell(row=idx, column=10).number_format = "#,##0.00"

    total_row = len(report.rows) + 4
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=float(report.total_purchase_amount))
    ws.cell(row=total_row, column=10, value=float(report.total_sales_amount))
    ws.cell(row=total_row, column=9).number_format = "#,##0.00"
    ws.cell(row=total_row, column=10).number_format = "#,##0.00"
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    ws.cell(row=total_row, column=10).font = Font(bold=True)

    widths = [16, 28, 16, 12, 12, 12, 12, 12, 14, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inventory-report-{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/salesperson-report", response_model=SalespersonReport)
def get_salesperson_report(
    year: int,
    month: int,
    db: DbDep,
    _: CurrentUser,
) -> SalespersonReport:
    try:
        return inventory_crud.salesperson_monthly_report(db, year=year, month=month)
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc


@router.get("/salesperson-report.xlsx")
def export_salesperson_report(
    year: int,
    month: int,
    db: DbDep,
    _: CurrentUser,
) -> StreamingResponse:
    try:
        report = inventory_crud.salesperson_monthly_report(db, year=year, month=month)
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc)) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    title = f"業務員銷售報表 {year}-{month:02d}"
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    headers = ["業務員", "角色", "訂單數", "總數量", "總金額", "佔比"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1976D2")
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_amount_f = float(report.total_amount) if report.total_amount else 0.0
    for idx, row in enumerate(report.rows, start=4):
        display = row.full_name or row.username
        ws.cell(row=idx, column=1, value=display)
        ws.cell(row=idx, column=2, value=row.role_name or "")
        ws.cell(row=idx, column=3, value=row.order_count)
        ws.cell(row=idx, column=4, value=row.total_qty)
        amt = float(row.total_amount)
        ws.cell(row=idx, column=5, value=amt).number_format = "#,##0.00"
        share = (amt / total_amount_f) if total_amount_f else 0
        ws.cell(row=idx, column=6, value=share).number_format = "0.0%"

    total_row = len(report.rows) + 4
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=report.total_order_count).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=report.total_qty).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=float(report.total_amount))
    ws.cell(row=total_row, column=5).number_format = "#,##0.00"
    ws.cell(row=total_row, column=5).font = Font(bold=True)

    widths = [22, 12, 12, 12, 16, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"salesperson-report-{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
